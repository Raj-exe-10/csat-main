from flask import Flask, render_template, request, jsonify,url_for,json
from mysql.connector import connect, Error
import logging
import jwt
from openpyxl import load_workbook
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer, SignatureExpired
import logging
logging.basicConfig(level=logging.DEBUG)
from datetime import datetime, timedelta

# Intialising the flask app 
app = Flask(__name__)


# Initialising the mail 
app.config['SECRET_KEY'] = '12b634b098308d0dd6a7472e6744a1d1'  # Use a real secret key in production
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'furynick830@gmail.com'
app.config['MAIL_PASSWORD'] = 'rhva curi vmfs ikcg'
mail = Mail(app)

# Serializer for generating tokens
s = URLSafeTimedSerializer(app.config['SECRET_KEY'])

# Database connection function
def get_db_connection():
    try:
        conn = connect(
            host='localhost',
            database='csat',
            user='root',
            password='Password@123'  # It's safer to use environment variables here!
        )
        return conn
    except Error as e:
        print("Failed to connect to database:", e)
        return None


# Route to display the main page
@app.route('/')
def index():
    conn = get_db_connection()
    if conn is None:
        return "Database connection failed. Please try again later.", 500
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM user_info")
        rows = cursor.fetchall()
        return render_template('index.html', data=rows)
    except Error as e:
        app.logger.error("Failed to fetch data:", e)
        return "Error fetching data. Please try again later.", 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


# API route to insert data into 'user_info' table
@app.route('/insert_user', methods=['POST'])
def insert_user():
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    data = request.get_json()
    if not all([data.get(field) for field in ['id', 'user_name', 'email', 'user_type', 'is_active']]):
        return jsonify({'error': 'Missing data'}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "Database connection failed"}), 500

    try:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO user_info (id, user_name, email, user_type, is_active) VALUES (%s, %s, %s, %s, %s)",
            (data['id'], data['user_name'], data['email'], data['user_type'], data['is_active'])
        )
        conn.commit()
        return jsonify({'message': 'Data inserted'}), 201
    except Error as e:
        conn.rollback()
        app.logger.error("Failed to insert data:", e)
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# API route to insert data into 'project_info' table
@app.route("/insert_project_info", methods=['POST'])  
def insert_project_info():
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    data = request.get_json()
    # Validate the necessary fields are in the data
    required_fields = ['account_id', 'project_name', 'project_started_on', 'project_end_date']
    if not all([data.get(field) for field in required_fields]):
        return jsonify({'error': 'Missing data for required fields'}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "Database connection failed"}), 500

    try:
        cursor = conn.cursor()
        # Adjust the SQL query to match your table and column names
        cursor.execute(
            "INSERT INTO project_info (account_id, project_name, project_started_on, project_end_date) VALUES (%s, %s, %s, %s)",
            (data['account_id'], data['project_name'], data['project_started_on'], data['project_end_date'])
        )
        conn.commit()
        return jsonify({'message': 'Project information inserted successfully'}), 201
    except Exception as e:
        conn.rollback()
        app.logger.error("Failed to insert project data:", e)
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# API route to insert data into 'account_info' table
@app.route("/insert_account_info", methods=['POST'])  
def insert_account_info():
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    data = request.get_json()
    # Updated the fields according to your new table structure
    required_fields = ['account_name', 'vertical_name', 'account_head_id']
    if not all([data.get(field) for field in required_fields]):
        return jsonify({'error': 'Missing data for required fields'}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "Database connection failed"}), 500

    try:
        cursor = conn.cursor()
        # Updated the SQL query to match your new table and column names
        cursor.execute(
            "INSERT INTO account_info (account_name, vertical_name, account_head_id) VALUES (%s, %s, %s)",
            (data['account_name'], data['vertical_name'], data['account_head_id'])
        )
        conn.commit()
        return jsonify({'message': 'Account information inserted successfully'}), 201
    except Exception as e:
        conn.rollback()
        app.logger.error("Failed to insert account data:", e)
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# API route to insert data into 'project_csat_request_info' table
@app.route("/insert_csat_request_info", methods=['POST'])  #
def insert_csat_request_info():
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    data = request.get_json()
    # Update the field checks according to the new table schema
    required_fields = ['account_id', 'project_id', 'csat_duration', 'customer_id']
    if not all([data.get(field) for field in required_fields]):
        return jsonify({'error': 'Missing data for required fields'}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "Database connection failed"}), 500

    try:
        cursor = conn.cursor()
        # Update the SQL query to insert data into the 'project_csat_request_info' table
        cursor.execute(
            "INSERT INTO project_csat_request_info (account_id, project_id, csat_duration, customer_id) VALUES (%s, %s, %s, %s)",
            (data['account_id'], data['project_id'], data['csat_duration'], data['customer_id'])
        )
        conn.commit()
        return jsonify({'message': 'CSAT request information inserted successfully'}), 201
    except Exception as e:
        conn.rollback()
        app.logger.error("Failed to insert CSAT request data:", e)
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# API route to insert data into 'project_csat_customer_rating' table
@app.route("/insert_csat_rating", methods=['POST'])  
def insert_csat_rating():
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    data = request.get_json()
    # Update the field checks to match your new table columns
    required_fields = ['csat_request_id', 'csat_question_id', 'csat_rating']
    if not all([data.get(field) for field in required_fields]):
        return jsonify({'error': 'Missing data for required fields'}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "Database connection failed"}), 500

    try:
        cursor = conn.cursor()
        # Update the SQL query to insert data into the 'project_csat_customer_rating' table
        cursor.execute(
            "INSERT INTO project_csat_customer_rating (csat_request_id, csat_question_id, csat_rating) VALUES (%s, %s, %s)",
            (data['csat_request_id'], data['csat_question_id'], data['csat_rating'])
        )
        conn.commit()
        return jsonify({'message': 'CSAT rating inserted successfully'}), 201
    except Exception as e:
        conn.rollback()
        app.logger.error("Failed to insert CSAT rating data:", e)
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# API route to insert data into 'project_csat_questions' table
@app.route("/insert_csat_question", methods=['POST'])  
def insert_csat_question():
    if not request.is_json:
        return jsonify({"error": "Request must be JSON"}), 400

    data = request.get_json()
    # Update the field checks to match your new table columns
    required_fields = ['csat_request_id', 'questions']
    if not all([data.get(field) for field in required_fields]):
        return jsonify({'error': 'Missing data for required fields'}), 400

    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "Database connection failed"}), 500

    try:
        cursor = conn.cursor()
        # Update the SQL query to insert data into the 'project_csat_customer_rating' table
        cursor.execute(
            "INSERT INTO project_csat_questions (csat_request_id, questions) VALUES (%s, %s)",
            ( data['csat_request_id'], data['questions'])
        )
        conn.commit()
        return jsonify({'message': 'CSAT question inserted successfully'}), 201
    except Exception as e:
        conn.rollback()
        app.logger.error("Failed to insert CSAT rating data:", e)
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


# Route to insert data into the project_csat_customer_rating table
@app.route('/insert_csat_rating', methods=['POST'])
def insert_csat_rating():
    # Obtain data from request
    data = request.get_json()
    csat_request_id = data.get('csat_request_id')
    csat_question_id = data.get('csat_question_id')
    csat_rating = data.get('csat_rating')

    # Check if all necessary data is present
    if not all([csat_request_id, csat_question_id, csat_rating]):
        return jsonify({"error": "Missing one or more required fields"}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "Database connection failed"}), 500

    try:
        cursor = conn.cursor()
        query = """
        INSERT INTO project_csat_customer_rating (csat_request_id, csat_question_id, csat_rating)
        VALUES (%s, %s, %s)
        """
        cursor.execute(query, (csat_request_id, csat_question_id, csat_rating))
        conn.commit()
        return jsonify({"message": "CSAT rating inserted successfully"}), 201
    except Error as e:
        conn.rollback()  # Roll back the transaction in case of error
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

            
# fetching the projects related to a particular account id.
@app.route("/get_projects_by_account_id/<int:account_id>", methods=["GET"])
def get_projects_by_account_id(account_id):
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "Database connection failed"}), 500

    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, project_name FROM project_info WHERE account_id = %s",
            (account_id,),
        )
        projects = cursor.fetchall()
        return jsonify({"projects": projects})
    except Exception as e:
        app.logger.error("Failed to get projects for account ID:", account_id, e)
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

#fetching all the accounts from the database
@app.route("/get_accounts", methods=["GET"])
def get_accounts():
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "Database connection failed"}), 500

    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id, account_name FROM account_info")
        accounts = cursor.fetchall()
        return jsonify({"accounts": accounts})
    except Exception as e:
        app.logger.error("Failed to get accounts:", e)
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

#fetching all csat_id from the database
@app.route("/get_csat_id", methods=["GET"])
def get_csat_id():
    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "Database connection failed"}), 500

    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id  FROM project_csat_request_info")
        csat_ids = cursor.fetchall()
        return jsonify({"csat_id": csat_ids})
    except Exception as e:
        app.logger.error("Failed to get accounts:", e)
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# api for uploading a questions from an excel sheet.
@app.route("/upload_questions", methods=["POST"])
def upload_questions():
  if 'questionFile' not in request.files:
    return jsonify({"error": "No file uploaded"}), 400

  file = request.files['questionFile']
  if not file.filename.endswith(".xlsx"):
    return jsonify({"error": "Invalid file format. Only .xlsx files allowed"}), 400

  try:
    workbook = load_workbook(file, data_only=True)  # Read only for security
    sheet = workbook.active

    questions = []
    for row in sheet.iter_rows(min_row=2):  # Skip the header row
      question_text = row[0].value
      questions.append(question_text)

    return jsonify({"questions": questions})
  except Exception as e:
    app.logger.error("Failed to process uploaded file:", e)
    return jsonify({"error": "Error processing file"}), 500


@app.route('/send_review_link', methods=['POST'])
def send_review_link():
    csat_id = request.json.get('csat_id')
    questions = request.json.get('questions')
    customer_id = request.json.get('customer_id')
    print('questions',questions)
    # Create a token that expires in 24 hours
    expiration = datetime.utcnow() + timedelta(hours=24)
    token = jwt.encode({
        'csat_id': csat_id,
        'questions': questions,
        'customer_id': customer_id,
        'exp': expiration
    }, app.config['SECRET_KEY'], algorithm='HS256')

    conn = get_db_connection()
    if conn is None:
        return jsonify({"error": "Database connection failed"}), 500

    try:
        cursor = conn.cursor()
        cursor.execute("SELECT email FROM user_info WHERE id = %s", (customer_id,))
        customer_email = cursor.fetchone()
        if customer_email is None:
            return jsonify({"error": "No email found for given customer ID"}), 404

        email = customer_email[0]  # Assuming the email is in the first column of the results
        link = url_for('review_user', token=token, _external=True)

        # Send the email
        msg = Message("Review Questions", sender=app.config['MAIL_USERNAME'], recipients=[email])
        msg.body = f"Please complete the following review questions: {link}"
        mail.send(msg)
        return jsonify({'message': 'Email sent successfully!'}), 200
    except Exception as e:
        app.logger.error("Failed to send email:", e)
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/review_user/<token>')
def review_user(token):
    try:
        # Decode the JWT token
        data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])

        # Extract the csat_id, questions, and customer_id from the token
        csat_id = data['csat_id']
        questions = data['questions']  # Questions are now coming directly from the token
        print(questions)
        customer_id = data['customer_id']  # Customer ID is also retrieved from the token

        # Render the HTML template with the data from the token
        return render_template('review_user.html', csat_id=csat_id, questions=questions, customer_id=customer_id)
    except jwt.ExpiredSignatureError:
        return 'This link has expired.', 403
    except Exception as e:
        return str(e), 500



# Route to serve the insertion form
@app.route('/insert_user_page')
def form():
    return render_template('insert_user.html')

# Route to serve the insertion form
@app.route('/insert_project_info_page')
def insert_project_info_page():
    return render_template('insert_project_info.html')

# Route to serve the insertion form
@app.route('/insert_account_info_page')
def insert_account_info_page():
    return render_template('insert_account_info.html')


@app.route('/upload_questions_page')
def upload_questions_page():
    csat_id = request.args.get('csat_id')  # Access the CSAT ID from the query parameter
    customer_id = request.args.get('customer_id') # Access the customer_id as the query parameter
    print(csat_id)
    print(customer_id)
    return render_template('questions.html', csat_id=csat_id,customer_id=customer_id)

# Route to page for uploading questions from the excel file on the local system
@app.route('/csat_id_page')
def csat_id_page():
    return render_template('csat_id.html')

@app.route('/review', methods=['GET'])
def review_questions():
    questions = json.loads(request.args.get('questions'))
    csat_id = request.args.get('csatId')
    return render_template('review.html', questions=questions, csat_id=csat_id)

@app.route('/submit_ratings_admin', methods=['POST'])
def submit_ratings_admin():
    ratings = request.get_json()
    # Process the ratings data as needed
    # For example, you can save the ratings to a database
    print(ratings)
    return jsonify({'success': True})

# Start the Flask app
if __name__ == '__main__':
    app.run(debug=True)